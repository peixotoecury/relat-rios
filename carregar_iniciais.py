# -*- coding: utf-8 -*-
"""
carregar_iniciais.py — carrega o primeiro mês de dado real dos 6 clientes
novos (Bridgestone já veio pela migração). Reaproveita a mesma lógica do
upload via navegador (index.html), mas em Python, direto dos arquivos que
a usuária mandou — evita ter que repetir 6x o upload manual pela tela.

Uso:
    python carregar_iniciais.py
"""
import sys
import re
from datetime import datetime

import openpyxl
import requests

sys.stdout.reconfigure(encoding="utf-8")

SUPABASE_URL = "https://rpibvjcnrseuugpkfmdj.supabase.co"
SUPABASE_KEY = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJwaWJ2amNucnNldXVncGtmbWRqIiwi"
                "cm9sZSI6ImFub24iLCJpYXQiOjE3ODE1NTc3MTcsImV4cCI6MjA5NzEzMzcxN30.ecihol8JESMH7cgFSvWKIzp-OwoPRFqdK3aCDwpCeg8")
HEADERS = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}", "Content-Type": "application/json"}


def num(v):
    if v is None or v == "" or v == "***" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").strip()
    if s.count(",") and s.count("."):
        s = s.replace(".", "").replace(",", ".")
    elif s.count(","):
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def data_iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None


SIGLAS = {"LTDA", "S.A.", "SA", "ME", "EPP", "EIRELI", "MEI", "CNPJ", "CPF", "INSS", "IRRF", "FGTS", "CLT", "PJE", "TRT", "TST"}
MINUSCULAS = {"de", "da", "do", "das", "dos", "e", "em", "a", "o"}


def padronizar(v):
    if not isinstance(v, str):
        return v
    s = re.sub(r"\s+", " ", v).strip()
    if not s or re.fullmatch(r"[\d./\-:,R$\s]+", s):
        return s
    palavras = []
    for p in s.split(" "):
        pu = p.upper().rstrip(".,")
        if pu in SIGLAS:
            palavras.append(p.upper())
        elif p.lower() in MINUSCULAS:
            palavras.append(p.lower())
        elif len(p) <= 1:
            palavras.append(p.upper())
        else:
            palavras.append(p[0].upper() + p[1:].lower())
    palavras[0] = palavras[0][0].upper() + palavras[0][1:]
    return " ".join(palavras)


def get_cliente_e_config(slug):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/rel_clientes", headers=HEADERS,
                      params={"slug": f"eq.{slug}", "select": "id"}, timeout=30)
    cliente_id = r.json()[0]["id"]
    r = requests.get(f"{SUPABASE_URL}/rest/v1/rel_config_cliente", headers=HEADERS,
                      params={"cliente_id": f"eq.{cliente_id}", "select": "config", "order": "id.desc", "limit": "1"}, timeout=30)
    return cliente_id, r.json()[0]["config"]


def ler_planilha(caminho, aba, header_row):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    header = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    idx = {}
    for i, h in enumerate(header):
        nome = str(h).strip() if h is not None else ""
        if nome and nome not in idx:
            idx[nome] = i
    linhas = [r for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True)
              if r and any(v is not None for v in r)]
    wb.close()
    return idx, linhas


def montar_dados(colunas, idx, linhas):
    chave_col = next(c for c in colunas if c.get("chave"))
    faltando = [c["origem"] for c in colunas if c["origem"] not in idx]
    if faltando:
        print(f"  [AVISO] colunas não encontradas na planilha: {faltando}")
    por_processo = {}
    for r in linhas:
        linha = {}
        for c in colunas:
            if c["origem"] not in idx:
                continue
            v = r[idx[c["origem"]]]
            if c.get("tipo") == "numero":
                v = num(v)
            elif isinstance(v, datetime):
                v = data_iso(v)
            else:
                v = padronizar(v)
            linha[c["origem"]] = v
        processo = linha.get(chave_col["origem"])
        if not processo:
            continue
        por_processo[str(processo).strip()] = linha
    return por_processo


def upsert_lote(tabela, registros, on_conflict, lote=250):
    for i in range(0, len(registros), lote):
        chunk = registros[i:i + lote]
        r = requests.post(f"{SUPABASE_URL}/rest/v1/{tabela}?on_conflict={on_conflict}",
                           headers={**HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"},
                           json=chunk, timeout=60)
        if not r.ok:
            print(f"  [ERRO] {r.status_code} {r.text[:400]}")
            r.raise_for_status()


def carregar_cliente(slug, caminho, aba, header_row, mes_referencia):
    print(f"=== {slug} ===")
    cliente_id, config = get_cliente_e_config(slug)
    idx, linhas = ler_planilha(caminho, aba, header_row)
    print(f"  {len(linhas)} linhas de dado na planilha.")
    por_processo = montar_dados(config["colunas_processos"], idx, linhas)
    print(f"  {len(por_processo)} processos com chave válida.")

    encerrado_cfg = config.get("encerrado", {"fonte": "comparacao_mes_anterior"})
    status_por_processo = {}
    if encerrado_cfg["fonte"] == "campo_status":
        vals_enc = {v.lower() for v in encerrado_cfg.get("valores_encerrado", [])}
        vals_novo = {v.lower() for v in encerrado_cfg.get("valores_novo", [])}
        for processo, linha in por_processo.items():
            v = str(linha.get(encerrado_cfg["campo"], "") or "").lower()
            status_por_processo[processo] = "encerrado" if v in vals_enc else ("novo" if v in vals_novo else "ativo")
    else:
        # primeiro mês do cliente -> tudo novo (não há mês anterior no schema novo)
        for processo in por_processo:
            status_por_processo[processo] = "novo"

    registros = [{
        "cliente_id": cliente_id, "mes_referencia": mes_referencia, "processo": processo,
        "dados": dados, "status_calculado": status_por_processo.get(processo, "ativo"),
    } for processo, dados in por_processo.items()]
    upsert_lote("rel_processos_snapshot", registros, "cliente_id,mes_referencia,processo")
    print(f"  {len(registros)} processos gravados em {mes_referencia}.")


CARGAS = [
    dict(slug="bridgestone-contingencia",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\Bridgestone\17. BASES PARA ELABORAÇÃO RELATÓRIOS\FORECAST 08_2026\Contingency\Labor - August 2026 - Vfinal.xlsx",
         aba="Ago.2026", header_row=3, mes_referencia="2026-08-01"),
    dict(slug="bridgestone-forecast",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\Bridgestone\17. BASES PARA ELABORAÇÃO RELATÓRIOS\FORECAST 08_2026\Forecast\Forecast - Jul.2026 - Vfinal i.xlsx",
         aba="Jun.2026", header_row=5, mes_referencia="2026-06-01"),
    dict(slug="cbc",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\CBC\08.2026\CBC - Relatório Contingência Trabalhista - Julho 2026 (1).xlsx",
         aba="Base P&C - Jul. 2026 - CBC", header_row=3, mes_referencia="2026-07-01"),
    dict(slug="embraer-reintegrados",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\Embraer\2026\Estratégicos e Reintegrados\09 - Setembro\Planilha de Reintegrados e Estratégicos - P&C - Setembro - 2026.xlsx",
         aba="Reintegrados", header_row=2, mes_referencia="2026-09-01"),
    dict(slug="embraer-estrategicos",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\Embraer\2026\Estratégicos e Reintegrados\09 - Setembro\Planilha de Reintegrados e Estratégicos - P&C - Setembro - 2026.xlsx",
         aba=None, header_row=1, mes_referencia="2026-09-01"),  # aba resolvida pelo índice abaixo
    dict(slug="ferrero",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\Ferrero\Relatórios\2026\08 - Agosto\Relatório Contingências Trabalhistas - Ferrero - Agosto.xlsx",
         aba="Base P&C - Agosto.2026", header_row=3, mes_referencia="2026-08-01"),
    dict(slug="t4f",
         caminho=r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Relatórios e Controles - Time B\T4F\2026\3Q2026\PC - Relatório de Provisões Trabalhista - T4F - 2Q2026 - VFinal II.xlsx",
         aba="2Q2026", header_row=3, mes_referencia="2026-06-01"),  # trimestral -- usando o mês final do 2Q como referência, AVISAR usuária
]


def main():
    for carga in CARGAS:
        aba = carga["aba"]
        if aba is None:  # Embraer - Estratégicos: nome da aba com encoding chato, resolve pelo índice
            wb = openpyxl.load_workbook(carga["caminho"], read_only=True)
            aba = wb.sheetnames[2]
            wb.close()
        carregar_cliente(carga["slug"], carga["caminho"], aba, carga["header_row"], carga["mes_referencia"])
    print("\nConcluído.")


if __name__ == "__main__":
    main()
